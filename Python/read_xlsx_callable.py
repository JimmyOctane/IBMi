#!/QOpenSys/pkgs/bin/python3
"""
Excel Reader for IBM i IFS - No Pandas Version
Uses only openpyxl and pyodbc
"""

import sys
import pyodbc
import json
from datetime import datetime
import traceback
from openpyxl import load_workbook

def log_message(message, log_file='/tmp/xlsx_import.log'):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(log_file, 'a') as f:
        f.write(f"[{timestamp}] {message}\n")

def read_xlsx_from_ifs(xlsx_path, sheet_name=0, start_row=0):
    try:
        log_message(f"Reading Excel file: {xlsx_path}")
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        sheet = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
        all_rows = list(sheet.iter_rows(values_only=True))
        headers = list(all_rows[start_row])
        data_rows = all_rows[start_row + 1:]
        log_message(f"Read {len(data_rows)} rows, Columns: {headers}")
        wb.close()
        return {'headers': headers, 'data': data_rows, 'row_count': len(data_rows)}
    except Exception as e:
        log_message(f"ERROR: {str(e)}")
        raise

def connect_db2(library):
    conn_string = f"DRIVER={{IBM i Access ODBC Driver}};SYSTEM=localhost;DBQ={library};CMT=0;"
    return pyodbc.connect(conn_string)

def analyze_column_type(data_rows, col_idx):
    values = [row[col_idx] for row in data_rows if col_idx < len(row) and row[col_idx] is not None]
    if not values:
        return "VARCHAR(50)"
    
    sample = values[0]
    if isinstance(sample, (int, float)):
        if all(isinstance(v, int) for v in values):
            max_val = max(abs(v) for v in values)
            if max_val < 32767:
                return "SMALLINT"
            elif max_val < 2147483647:
                return "INTEGER"
            return "BIGINT"
        else:
            max_val = max(abs(v) for v in values if isinstance(v, (int, float)))
            int_part = len(str(int(max_val))) if max_val > 0 else 1
            if col_idx == 3 or col_idx == 4:
                return f"DECIMAL({max(int_part + 2, 10)}, 2)"
            return f"DECIMAL({max(int_part + 5, 15)}, 5)"
    elif isinstance(sample, datetime):
        return "TIMESTAMP"
    else:
        max_len = max(len(str(v)) for v in values)
        max_len = int(max_len * 1.2)
        max_len = max(min(max_len, 1000), 10)
        return f"VARCHAR({max_len})"

def create_or_replace_table(excel_data, conn, library, table):
    cursor = conn.cursor()
    table_full = f"{library}.{table}"
    try:
        cursor.execute(f"DROP TABLE {table_full}")
        conn.commit()
    except:
        pass
    
    col_defs = []
    for i, col_name in enumerate(excel_data['headers']):
        col_name_clean = str(col_name).replace(' ', '_').replace('-', '_').upper() or f"COL{i+1}"
        db_type = analyze_column_type(excel_data['data'], i)
        col_defs.append(f"{col_name_clean} {db_type}")
        log_message(f"Column {i+1} ({col_name_clean}): {db_type}")
    
    create_sql = f"CREATE TABLE {table_full} ({', '.join(col_defs)})"
    cursor.execute(create_sql)
    conn.commit()
    cursor.close()

def import_to_table(excel_data, conn, library, table):
    cursor = conn.cursor()
    col_count = len(excel_data['headers'])
    placeholders = ', '.join(['?' for _ in range(col_count)])
    insert_sql = f"INSERT INTO {library}.{table} VALUES ({placeholders})"
    
    success_count = 0
    error_count = 0
    
    for row in excel_data['data']:
        try:
            values = []
            for col_idx in range(col_count):
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    values.append(None)
                elif col_idx == 3 or col_idx == 4:
                    try:
                        values.append(round(float(val), 2))
                    except:
                        values.append(val)
                else:
                    values.append(val)
            cursor.execute(insert_sql, values)
            success_count += 1
        except Exception as e:
            error_count += 1
            log_message(f"ERROR on row: {str(e)}")
    
    conn.commit()
    cursor.close()
    return success_count, error_count

def main():
    try:
        if len(sys.argv) < 4:
            print("Usage: read_xlsx_callable.py <xlsx_path> <library> <table> [sheet] [start_row]")
            sys.exit(1)
        
        xlsx_path = sys.argv[1]
        library = sys.argv[2].upper()
        table = sys.argv[3].upper()
        sheet_name = int(sys.argv[4]) if len(sys.argv) > 4 else 0
        start_row = int(sys.argv[5]) if len(sys.argv) > 5 else 0
        
        log_message("="*60)
        log_message(f"File: {xlsx_path}, Target: {library}.{table}")
        
        excel_data = read_xlsx_from_ifs(xlsx_path, sheet_name, start_row)
        conn = connect_db2(library)
        create_or_replace_table(excel_data, conn, library, table)
        success, errors = import_to_table(excel_data, conn, library, table)
        conn.close()
        
        summary = {"status": "success", "rows_imported": success, "rows_failed": errors}
        print(json.dumps(summary))
        log_message(f"Complete: {json.dumps(summary)}")
        sys.exit(0 if errors == 0 else 1)
    except Exception as e:
        error_msg = {"status": "error", "message": str(e)}
        print(json.dumps(error_msg))
        log_message(f"FATAL: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
